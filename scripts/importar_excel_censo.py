#!/usr/bin/env python3
"""Importa un Excel de censo externo con verificación Nexus y seguridad.

Uso típico:

  NEXUS_SCRIPT_EMAIL=admin@refugio.app NEXUS_SCRIPT_PASSWORD=... \\
    python3 scripts/importar_excel_censo.py \\
      --archivo /tmp/CE_Andres_Bello.xlsx --centro-id centro-32 --con-nexus --dry-run

  NEXUS_SCRIPT_EMAIL=admin@refugio.app NEXUS_SCRIPT_PASSWORD=... \\
    python3 scripts/importar_excel_censo.py \\
      --archivo /tmp/CE_Andres_Bello.xlsx --centro-id centro-32 --con-nexus --aplicar
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import time
import unicodedata
import urllib.error
import urllib.request
from collections import Counter
from concurrent.futures import FIRST_COMPLETED, Future, ThreadPoolExecutor, wait
from dataclasses import dataclass
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
ENV = ROOT / ".env"
DEFAULT_URL = "https://xzwifkckkakldnzkdeby.supabase.co"
DEFAULT_GATEWAY = "https://nexus.m0n1t0r-d3-3v3nt0s.net"
DEFAULT_CONCURRENCY = 5
DEFAULT_RATE = 2.5
DEFAULT_CIRCUIT_BREAKER = 5
DEFAULT_NEXUS_TIMEOUT = 20.0
PROGRESS_INTERVAL = 10.0
PAGE_SIZE = 1000
SIN_CEDULA = {"", "S/C", "SC", "S/N", "SN", "S/D", "SD", "SIN CEDULA", "SIN CÉDULA", "N/A", "NA", "-"}
LETRAS_NEXUS = {"V", "E"}

ALIASES_NOMBRE: dict[str, str] = {
    "delgado chalbaud": "centro-33",
    "delgado chalboud": "centro-33",
    "perez bonalde": "centro-34",
    "pérez bonalde": "centro-34",
    "bonalde": "centro-34",
    "lossada": "centro-11",
    "losada": "centro-11",
    "mama rosa": "centro-36",
    "mamá rosa": "centro-36",
}

PREFIJOS = re.compile(
    r"^(u\.?\s*e\.?\s*n\.?\s*b\.?|u\.?\s*e\.?\s*n\.?|u\.?\s*e\.?\s*d\.?|u\.?\s*e\.?\s*e\.?|"
    r"u\.?\s*e\.?\s*|e\.?\s*n\.?\s*b\.?|e\.?\s*b\.?\s*e\.?|e\.?\s*b\.?\s*n\.?|"
    r"e\.?\s*t\.?\s*i\.?|c\.?\s*e\.?\s*i\.?\s*n\.?|c\.?\s*e\.?\s*i\.?|c\.?\s*e\.?\s*n\.?|"
    r"c\.?\s*e\.?\s*|complejo\s+educativo|liceo|escuela(\s+integral\s+basica)?|"
    r"unidad\s+educativa|refugio(\s+para)?|universidad|polideportivo|estadio|"
    r"fundacion|campamento|centro\s+de\s+educacion\s+inicial)\s+",
    re.I,
)


@dataclass
class CentroApp:
    id: str
    nombre: str
    activo: bool


def strip_accents(texto: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", texto or "") if unicodedata.category(c) != "Mn"
    )


def key(texto: str) -> str:
    texto = strip_accents(str(texto or "").lower())
    texto = re.sub(r"[^a-z0-9]+", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


def normalizar_nombre_centro(texto: str) -> str:
    texto = strip_accents((texto or "").lower()).replace("ü", "u")
    texto = PREFIJOS.sub("", texto)
    texto = re.sub(r"[^a-z0-9\s]", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


def tokens(texto: str) -> set[str]:
    return {t for t in normalizar_nombre_centro(texto).split() if len(t) > 1}


def score_nombres(a: str, b: str) -> float:
    na, nb = normalizar_nombre_centro(a), normalizar_nombre_centro(b)
    if not na or not nb:
        return 0.0
    if na == nb:
        return 1.0
    if na in nb or nb in na:
        return 0.92
    ta, tb = tokens(a), tokens(b)
    if not ta or not tb:
        return 0.0
    inter = len(ta & tb)
    union = len(ta | tb)
    jaccard = inter / union
    nucleo_a = na.split()[-1] if na.split() else ""
    nucleo_b = nb.split()[-1] if nb.split() else ""
    if nucleo_a and nucleo_a == nucleo_b and inter >= 1:
        jaccard = max(jaccard, 0.75)
    return jaccard


def cargar_env() -> tuple[str, str, str]:
    valores: dict[str, str] = {}
    claves = (
        "VITE_SUPABASE_URL",
        "VITE_SUPABASE_ANON_KEY",
        "VITE_NEXUS_GATEWAY_URL",
        "NEXUS_SCRIPT_EMAIL",
        "NEXUS_SCRIPT_PASSWORD",
    )
    if ENV.exists():
        for linea in ENV.read_text(encoding="utf-8").splitlines():
            linea = linea.strip()
            for clave in claves:
                if linea.startswith(f"{clave}="):
                    valores[clave] = linea.split("=", 1)[1].strip().strip('"').strip("'")
    # Credenciales de script: .env basta; env del shell gana si ya están seteadas.
    for clave in ("NEXUS_SCRIPT_EMAIL", "NEXUS_SCRIPT_PASSWORD"):
        if valores.get(clave) and not os.environ.get(clave):
            os.environ[clave] = valores[clave]
    url = os.environ.get("VITE_SUPABASE_URL", valores.get("VITE_SUPABASE_URL", DEFAULT_URL))
    anon = os.environ.get("VITE_SUPABASE_ANON_KEY", valores.get("VITE_SUPABASE_ANON_KEY", ""))
    gateway = os.environ.get(
        "VITE_NEXUS_GATEWAY_URL",
        valores.get("VITE_NEXUS_GATEWAY_URL", DEFAULT_GATEWAY),
    )
    if not anon:
        raise SystemExit("Falta VITE_SUPABASE_ANON_KEY (.env o variable de entorno)")
    return url.rstrip("/"), anon, gateway.rstrip("/")


def autenticar(url: str, anon_key: str) -> str:
    email = os.environ.get("NEXUS_SCRIPT_EMAIL")
    password = os.environ.get("NEXUS_SCRIPT_PASSWORD")
    if not email or not password:
        raise SystemExit(
            "Faltan NEXUS_SCRIPT_EMAIL / NEXUS_SCRIPT_PASSWORD (.env o variables de entorno)"
        )
    req = urllib.request.Request(
        f"{url}/auth/v1/token?grant_type=password",
        data=json.dumps({"email": email, "password": password}).encode("utf-8"),
        headers={"apikey": anon_key, "Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            body = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        detalle = exc.read().decode("utf-8", errors="replace")
        raise SystemExit(f"No se pudo autenticar ({exc.code}): {detalle}") from exc
    token = body.get("access_token")
    if not token:
        raise SystemExit(f"Respuesta de login sin access_token: {body}")
    return token


def rpc(url: str, anon_key: str, jwt: str, fn: str, payload: dict[str, Any]) -> object:
    req = urllib.request.Request(
        f"{url}/rest/v1/rpc/{fn}",
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "apikey": anon_key,
            "Authorization": f"Bearer {jwt}",
            "Content-Type": "application/json",
            "Prefer": "return=representation",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=180) as resp:
            raw = resp.read().decode("utf-8")
            return json.loads(raw) if raw else None
    except urllib.error.HTTPError as exc:
        detalle = exc.read().decode("utf-8", errors="replace")
        raise SystemExit(f"RPC {fn} falló HTTP {exc.code}: {detalle}") from exc


def rest_get_paginado(url: str, anon_key: str, jwt: str, path: str) -> list[dict[str, Any]]:
    """Lee todos los resultados PostgREST respetando límite por página."""
    acumulado: list[dict[str, Any]] = []
    desde = 0
    while True:
        hasta = desde + PAGE_SIZE - 1
        req = urllib.request.Request(
            f"{url}/rest/v1/{path}",
            headers={
                "apikey": anon_key,
                "Authorization": f"Bearer {jwt}",
                "Range-Unit": "items",
                "Range": f"{desde}-{hasta}",
            },
        )
        with urllib.request.urlopen(req, timeout=60) as resp:
            pagina = json.loads(resp.read().decode("utf-8"))
        if not isinstance(pagina, list):
            raise SystemExit(f"Respuesta REST inesperada para {path}")
        acumulado.extend(pagina)
        if len(pagina) < PAGE_SIZE:
            return acumulado
        desde += PAGE_SIZE


def cargar_cache_nexus(
    url: str,
    anon_key: str,
    jwt: str,
) -> dict[tuple[str, str], dict[str, Any]]:
    """Carga verificaciones persistentes. Una clave presente jamás toca Nexus."""
    filas = rest_get_paginado(
        url,
        anon_key,
        jwt,
        "nexus_consultas?select=letra,cedula,data,actualizado_ts&order=letra,cedula",
    )
    cache: dict[tuple[str, str], dict[str, Any]] = {}
    for fila in filas:
        letra = texto(fila.get("letra")).upper()
        cedula = texto(fila.get("cedula"))
        data = fila.get("data")
        if letra in LETRAS_NEXUS and cedula and isinstance(data, dict) and data.get("ok") is not False:
            cache[(letra, cedula)] = data
    return cache


def guardar_cache_nexus(
    url: str,
    anon_key: str,
    jwt: str,
    fichas: dict[tuple[str, str], dict[str, Any]],
) -> None:
    """Persiste fichas nuevas en lotes; próximas corridas no consultan Nexus."""
    if not fichas:
        return
    filas = [
        {
            "letra": letra,
            "cedula": cedula,
            "data": data,
            "actualizado_ts": int(time.time() * 1000),
            "actualizado_por": "script:importar_excel_censo",
        }
        for (letra, cedula), data in fichas.items()
    ]
    for inicio in range(0, len(filas), 200):
        req = urllib.request.Request(
            f"{url}/rest/v1/nexus_consultas?on_conflict=letra,cedula",
            data=json.dumps(filas[inicio : inicio + 200]).encode("utf-8"),
            headers={
                "apikey": anon_key,
                "Authorization": f"Bearer {jwt}",
                "Content-Type": "application/json",
                "Prefer": "resolution=merge-duplicates,return=minimal",
            },
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=60):
            pass


def listar_centros(url: str, anon_key: str, jwt: str) -> list[CentroApp]:
    req = urllib.request.Request(
        f"{url}/rest/v1/centros?select=id,data,deleted&limit=500",
        headers={"apikey": anon_key, "Authorization": f"Bearer {jwt}"},
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        rows = json.loads(resp.read().decode("utf-8"))
    return [
        CentroApp(
            id=row["id"],
            nombre=((row.get("data") or {}).get("nombre") or row["id"]).strip(),
            activo=not bool(row.get("deleted")),
        )
        for row in rows
    ]


def consultar_nexus(
    gateway: str,
    jwt: str,
    letra: str,
    cedula: str,
    timeout: float,
) -> dict[str, Any]:
    req = urllib.request.Request(
        f"{gateway}/v1/person/search/external/full/{letra}/{cedula}/censo",
        data=b"{}",
        headers={"Authorization": f"Bearer {jwt}", "Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return {"status": resp.status, "body": json.loads(resp.read().decode("utf-8"))}
    except urllib.error.HTTPError as exc:
        raw = exc.read().decode("utf-8", errors="replace")
        try:
            body = json.loads(raw)
        except json.JSONDecodeError:
            body = {"error": raw}
        return {"status": exc.code, "body": body}
    except urllib.error.URLError as exc:
        return {"status": None, "body": {"error": str(exc.reason)}}


def consultar_nexus_concurrente(
    gateway: str,
    jwt: str,
    pendientes: list[tuple[str, str]],
    concurrency: int,
    rate: float,
    circuit_breaker: int,
    timeout: float,
) -> tuple[
    dict[tuple[str, str], dict[str, Any]],
    list[dict[str, Any]],
    int,
]:
    """Consulta claves únicas con límite calibrado y corte ante caída.

    Con concurrencia >1, envía una solicitud cada 1/N segundos y mantiene como
    máximo N en vuelo. Concurrencia 1 conserva modo secuencial con ``rate``.
    """
    fichas: dict[tuple[str, str], dict[str, Any]] = {}
    errores: list[dict[str, Any]] = []
    omitidas = 0
    fallos_consecutivos = 0
    procesadas = 0
    total = len(pendientes)
    ultimo_progreso = time.monotonic()

    def procesar(clave: tuple[str, str]) -> tuple[tuple[str, str], dict[str, Any]]:
        letra, cedula = clave
        return clave, consultar_nexus(gateway, jwt, letra, cedula, timeout)

    def registrar(
        clave: tuple[str, str],
        resultado: dict[str, Any],
    ) -> bool:
        nonlocal fallos_consecutivos, procesadas, ultimo_progreso
        procesadas += 1
        status = resultado.get("status")
        body = resultado.get("body")
        if status == 200 and isinstance(body, dict) and body.get("ok") is not False:
            fichas[clave] = body
            fallos_consecutivos = 0
        else:
            if status == 404 or (status == 200 and isinstance(body, dict) and body.get("ok") is False):
                fallos_consecutivos = 0
            else:
                fallos_consecutivos += 1
            errores.append(
                {
                    "tipo_doc": clave[0],
                    "documento": clave[1],
                    "status": status,
                    "detalle": body,
                }
            )
        ahora = time.monotonic()
        if procesadas == total or procesadas % 10 == 0 or ahora - ultimo_progreso >= PROGRESS_INTERVAL:
            print(
                f"Nexus: {procesadas}/{total} procesadas · "
                f"{len(fichas)} verificadas · {len(errores)} errores",
                file=sys.stderr,
                flush=True,
            )
            ultimo_progreso = ahora
        return fallos_consecutivos >= circuit_breaker

    if concurrency <= 1:
        for indice, clave in enumerate(pendientes):
            if indice > 0:
                time.sleep(rate)
            _, resultado = procesar(clave)
            if registrar(clave, resultado):
                omitidas = len(pendientes) - indice - 1
                break
        return fichas, errores, omitidas

    intervalo = 1.0 / concurrency
    proximas = iter(pendientes)
    en_vuelo: dict[Future[tuple[tuple[str, str], dict[str, Any]]], tuple[str, str]] = {}
    abortado = False
    ultimo_envio: float | None = None

    with ThreadPoolExecutor(max_workers=concurrency) as executor:
        while not abortado:
            while len(en_vuelo) < concurrency:
                try:
                    clave = next(proximas)
                except StopIteration:
                    break
                if ultimo_envio is not None:
                    espera = intervalo - (time.monotonic() - ultimo_envio)
                    if espera > 0:
                        time.sleep(espera)
                futuro = executor.submit(procesar, clave)
                ultimo_envio = time.monotonic()
                en_vuelo[futuro] = clave

            if not en_vuelo:
                break

            terminados, _ = wait(en_vuelo, timeout=1, return_when=FIRST_COMPLETED)
            if not terminados:
                ahora = time.monotonic()
                if ahora - ultimo_progreso >= PROGRESS_INTERVAL:
                    print(
                        f"Nexus: {procesadas}/{total} procesadas · "
                        f"{len(en_vuelo)} en vuelo (timeout {timeout:g}s)",
                        file=sys.stderr,
                        flush=True,
                    )
                    ultimo_progreso = ahora
                continue
            for futuro in terminados:
                clave = en_vuelo.pop(futuro)
                try:
                    _, resultado = futuro.result()
                except Exception as exc:  # noqa: BLE001
                    resultado = {"status": None, "body": {"error": str(exc)}}
                if registrar(clave, resultado):
                    abortado = True
                    break

        if abortado:
            omitidas = sum(1 for _ in proximas)
            for futuro, clave in list(en_vuelo.items()):
                try:
                    _, resultado = futuro.result()
                except Exception as exc:  # noqa: BLE001
                    resultado = {"status": None, "body": {"error": str(exc)}}
                registrar(clave, resultado)

    return fichas, errores, omitidas


def resolver_centro(nombre_raw: str, centros: list[CentroApp], forzado: str | None) -> tuple[str | None, str, str]:
    if forzado:
        centro = next((c for c in centros if c.id == forzado), None)
        if centro is None:
            return None, nombre_raw, "inexistente"
        if not centro.activo:
            return None, nombre_raw, "inactivo"
        return forzado, nombre_raw, "forzado"
    raw = (nombre_raw or "").strip()
    if not raw:
        return None, "", ""
    nombre_key = normalizar_nombre_centro(raw)
    if nombre_key in ALIASES_NOMBRE:
        centro_id = ALIASES_NOMBRE[nombre_key]
        centro = next((c for c in centros if c.id == centro_id), None)
        if centro is None:
            return None, raw, "inexistente"
        if not centro.activo:
            return None, raw, "inactivo"
        return centro_id, raw, "alias"
    mejor: CentroApp | None = None
    mejor_score = 0.0
    for centro in centros:
        puntaje = score_nombres(raw, centro.nombre)
        if puntaje > mejor_score:
            mejor_score = puntaje
            mejor = centro
    if mejor and not mejor.activo and mejor_score >= 0.75:
        return None, raw, "inactivo"
    if mejor and mejor_score >= 0.99:
        return mejor.id, raw, "exacto"
    if mejor and mejor_score >= 0.75:
        return mejor.id, raw, "fuzzy"
    return None, raw, ""


def _es_fila_header(celdas: list[str]) -> bool:
    """Detecta fila de encabezados reales (salta título mergeado arriba)."""
    unidos = " ".join(key(c) for c in celdas if c)
    marcadores = ("primer nombre", "documento", "cedula", "cédula", "apellido", "campamento")
    hits = sum(1 for m in marcadores if m in unidos)
    return hits >= 2


def leer_filas(path: Path) -> tuple[list[dict[str, Any]], str, list[str]]:
    sufijo = path.suffix.lower()
    if sufijo in {".xlsx", ".xlsm"}:
        try:
            import openpyxl  # type: ignore
        except ImportError as exc:
            raise SystemExit("Para .xlsx instale openpyxl: pip install openpyxl") from exc
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = None
        rows: list[tuple[Any, ...]] = []
        header_idx = -1
        for candidata in wb.worksheets:
            candidatas = list(candidata.iter_rows(min_row=1, max_row=20, values_only=True))
            for i, row in enumerate(candidatas):
                celdas = [str(h or "").strip() for h in row]
                if _es_fila_header(celdas):
                    ws = candidata
                    header_idx = i
                    break
            if ws is not None:
                break
        if ws is None:
            raise SystemExit("No se encontró una hoja con encabezados de censo")
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h or "").strip() for h in rows[header_idx]]
        salida: list[dict[str, Any]] = []
        for row in rows[header_idx + 1 :]:
            item = {}
            for i, header in enumerate(headers):
                if not header:
                    continue
                item[header] = row[i] if i < len(row) else None
            if any(v not in (None, "") for v in item.values()):
                salida.append(item)
        return salida, ws.title, headers

    text = path.read_text(encoding="utf-8-sig")
    dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
    reader = csv.DictReader(text.splitlines(), dialect=dialect)
    headers = [str(h or "").strip() for h in (reader.fieldnames or [])]
    return [{k.strip(): v for k, v in row.items() if k} for row in reader], path.name, headers


def pick(row: dict[str, Any], *aliases: str) -> Any:
    mapa = {key(k): v for k, v in row.items()}
    for alias in aliases:
        valor = mapa.get(key(alias))
        if valor not in (None, ""):
            return valor
    return ""


def texto(valor: Any) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def parse_bool(valor: Any) -> bool:
    if isinstance(valor, bool):
        return valor
    if isinstance(valor, (int, float)):
        return valor != 0
    limpio = key(texto(valor))
    return limpio in {"si", "sí", "s", "true", "t", "1", "yes", "y", "positivo", "x"}


def inferir_flags_seguridad(obs: str) -> dict[str, bool]:
    """Deriva flags desde texto libre SIIPOL/contrainteligencia."""
    t = key(obs)
    return {
        "registro_policial": (
            "posee registros policiales" in t
            or "posee registro policial" in t
            or "registros policiales" in t
        ),
        # Filosofía operativa / SIIPOL: denuncia de persona extraviada
        # = búsqueda activa = solicitada (misma bandeja KPI Solicitados).
        "solicitado": (
            "se encuentra solicitado" in t
            or "solicitado por" in t
            or "persona extraviada" in t
            or "registra como persona extraviada" in t
            or "extraviada" in t
            or "extraviado" in t
        ),
        "firmo_contra_presidente": (
            "firmo contra el presidente" in t or "firmo contra presidente" in t
        ),
        "deportado": "deportado" in t,
    }


def parse_cedula(raw: Any) -> tuple[str, str]:
    valor = texto(raw).upper()
    if valor in SIN_CEDULA:
        return "", ""
    match = re.match(r"^([VEP])[-\s.]?(\d+)$", valor)
    if match:
        return match.group(1), match.group(2)
    # Núcleos familiares: menores sin cédula como V-12345678-1 (padre + índice).
    # No son documento real → sin cédula (evitar concatenar dígitos falsos).
    if re.match(r"^([VEP])[-\s.]?\d+-\d+$", valor):
        return "", ""
    digitos = re.sub(r"\D", "", valor)
    if digitos and len(digitos) >= 5:
        return "V", digitos
    return "", ""


def split_nombre(nombre: str) -> tuple[str, str]:
    partes = nombre.strip().split(None, 1)
    if not partes:
        return "", ""
    if len(partes) == 1:
        return partes[0], ""
    return partes[0], partes[1]


def split_nombre_completo(nombre: str) -> tuple[str, str, str, str]:
    """Separa nombre completo cuando planilla no trae columnas individuales."""
    partes = nombre.strip().split()
    if len(partes) < 2:
        return (partes[0] if partes else ""), "", "", ""
    if len(partes) == 2:
        return partes[0], "", partes[1], ""
    if len(partes) == 3:
        return partes[0], partes[1], partes[2], ""
    return partes[0], partes[1], partes[2], " ".join(partes[3:])


def normalizar_sexo(valor: Any) -> str:
    limpio = key(texto(valor))
    if limpio in {"m", "masculino", "h", "hombre"}:
        return "M"
    if limpio in {"f", "femenino", "femenina", "mujer"}:
        return "F"
    return ""


def parse_edad(valor: Any) -> int | None:
    raw = texto(valor)
    if not raw:
        return None
    match = re.search(r"\d+", raw)
    if not match:
        return None
    edad = int(match.group(0))
    return edad if 0 <= edad <= 120 else None


def edad_nexus(body: dict[str, Any]) -> int | None:
    edad = body.get("edad")
    return int(edad) if isinstance(edad, int) and 0 <= edad <= 120 else None


def aplicar_nexus(payload: dict[str, Any], body: dict[str, Any], fuente: str = "nexus") -> None:
    payload["primer_nombre"] = texto(body.get("primer_nombre")) or payload["primer_nombre"]
    payload["segundo_nombre"] = texto(body.get("segundo_nombre")) or payload["segundo_nombre"]
    payload["primer_apellido"] = texto(body.get("primer_apellido")) or payload["primer_apellido"]
    payload["segundo_apellido"] = texto(body.get("segundo_apellido")) or payload["segundo_apellido"]
    payload["edad"] = "" if edad_nexus(body) is None else str(edad_nexus(body))
    sexo = normalizar_sexo(body.get("sexo"))
    if sexo:
        payload["sexo"] = sexo
    telefonos = body.get("telefonos")
    if not payload.get("telefono") and isinstance(telefonos, list) and telefonos:
        payload["telefono"] = texto(telefonos[0])
    payload["verificado_nexus"] = True
    payload["verificado_nexus_fuente"] = fuente if fuente in {"nexus", "cache"} else "nexus"


def fila_a_payload(
    row: dict[str, Any],
    centros: list[CentroApp],
    centro_forzado: str | None,
    col_centro: str | None,
) -> tuple[dict[str, Any] | None, dict[str, Any] | None]:
    ced_raw = pick(row, "cedula", "cédula", "documento", "ci", "doc")
    tipo_doc, documento = parse_cedula(ced_raw)
    tipo_doc_col = texto(pick(row, "tipo doc.", "tipo doc", "tipo_doc", "tipo documento")).upper()
    if tipo_doc_col in {"V", "E", "P"} and documento:
        tipo_doc = tipo_doc_col

    nombre = texto(pick(row, "nombres", "nombre", "primer nombre", "primer_nombre"))
    segundo_nombre = texto(pick(row, "segundo nombre", "segundo_nombre"))
    apellido = texto(pick(row, "apellidos", "apellido", "primer apellido", "primer_apellido"))
    segundo_apellido = texto(pick(row, "segundo apellido", "segundo_apellido"))
    if not nombre and not apellido:
        completo = texto(
            pick(
                row,
                "nombre completo",
                "nombre_completo",
                "nombre y apellido",
                "nombres y apellidos",
                "nombre y apellidos",
                "habitante nombres y apellidos",
                "habitante (nombres y apellidos)",
                "beneficiario",
                "persona",
            )
        )
        primer_nombre, segundo_nombre, primer_apellido, segundo_apellido = split_nombre_completo(completo)
    else:
        primer_nombre, resto_nombre = split_nombre(nombre)
        primer_apellido, resto_apellido = split_nombre(apellido)
        segundo_nombre = segundo_nombre or resto_nombre
        segundo_apellido = segundo_apellido or resto_apellido

    nombre_centro = texto(row.get(col_centro, "")) if col_centro else ""
    if not nombre_centro:
        nombre_centro = texto(pick(row, "campamento", "centro", "escuela", "refugio", "institucion", "institución"))
    centro_id, centro_raw, match = resolver_centro(nombre_centro, centros, centro_forzado)
    if not centro_id:
        return None, {
            "error": "centro_inactivo" if match == "inactivo" else "sin_centro",
            "nombre_centro_raw": centro_raw or nombre_centro,
        }

    edad = parse_edad(pick(row, "edad", "age"))
    tipo_registro = texto(pick(row, "tipo de registro", "tipo registro", "tipo registro policial"))
    descripcion_verificacion = texto(
        pick(
            row,
            "descripcion verificacion",
            "descripción verificación",
            "descripcion (verificacion)",
            "descripción (verificación)",
        )
    )
    observaciones_generales = texto(
        pick(
            row,
            "observaciones",
            "observacion",
            "observación",
            "informacion de interes",
            "información de interés",
            "sistemas de contrainteligencia y siipol",
            "sistemas de contrainteligencia",
            "siipol",
        )
    )
    observaciones = descripcion_verificacion or observaciones_generales
    flags_texto = inferir_flags_seguridad(observaciones)
    registro_policial = parse_bool(
        pick(
            row,
            "tiene registro policial",
            "registro policial",
            "reg policial",
            "reg. policial",
            "con reg policial",
            "con registro policial",
        )
    ) or flags_texto["registro_policial"]
    solicitado = parse_bool(
        pick(row, "está solicitado", "esta solicitado", "solicitado", "requerido")
    ) or flags_texto["solicitado"]
    firmo_contra_presidente = parse_bool(
        pick(
            row,
            "firmó contra presidente",
            "firmo contra presidente",
            "firmo vs pres",
            "firmo vs presidente",
            "firmó contra el gob.",
            "firmo contra el gob",
        )
    ) or flags_texto["firmo_contra_presidente"]
    deportado = parse_bool(pick(row, "deportado")) or flags_texto["deportado"]
    verificado_siipol = bool(descripcion_verificacion) or any(
        (registro_policial, solicitado, firmo_contra_presidente, deportado, bool(tipo_registro))
    )

    payload: dict[str, Any] = {
        "centro_id": centro_id,
        "nombre_centro_raw": centro_raw or nombre_centro,
        "centro_match": match,
        "primer_nombre": primer_nombre,
        "segundo_nombre": segundo_nombre,
        "primer_apellido": primer_apellido,
        "segundo_apellido": segundo_apellido,
        "edad": "" if edad is None else str(edad),
        "tipo_doc": tipo_doc,
        "documento": documento,
        "sexo": normalizar_sexo(pick(row, "sexo", "genero", "género")),
        "telefono": texto(
            pick(
                row,
                "telefono",
                "teléfono",
                "telefono principal",
                "teléfono principal",
                "celular",
                "phone",
            )
        ),
        "embarazada": parse_bool(pick(row, "embarazada", "embarazo")),
        "discapacidad": parse_bool(pick(row, "discapacidad", "discapacitado")),
        "discapacidad_detalle": texto(pick(row, "discapacidad detalle", "detalle discapacidad")),
        "enfermedad": parse_bool(pick(row, "enfermedad", "patologia", "patología")),
        "enfermedad_detalle": texto(pick(row, "enfermedad detalle", "detalle enfermedad", "patologia detalle")),
        "pais": texto(pick(row, "pais", "país")) or "Venezuela",
        "estado_federativo": texto(pick(row, "estado", "estado_federativo")),
        "municipio": texto(pick(row, "municipio")),
        "parroquia": texto(pick(row, "parroquia")),
        "calle": texto(pick(row, "direccion", "dirección", "calle", "sector")),
        "casa_edificio": texto(
            pick(
                row,
                "casa",
                "edificio",
                "casa_edificio",
                "aula",
                "ubicacion bloque carpa",
                "ubicacion (bloque/carpa)",
                "ubicacion",
                "ubicación",
            )
        ),
        "registro_policial": registro_policial,
        "solicitado": solicitado,
        "firmo_contra_presidente": firmo_contra_presidente,
        "deportado": deportado,
        "tipo_registro_policial": tipo_registro,
        "observaciones_seguridad": observaciones,
        "verificado_siipol": verificado_siipol,
        "verificado_nexus": False,
        "verificado_nexus_fuente": "",
    }
    return payload, None


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--archivo", required=True, type=Path)
    ap.add_argument("--centro-id", default=None, help="Fuerza centro_id si el Excel es de un solo campamento")
    ap.add_argument("--col-centro", default=None, help="Columna con nombre del campamento")
    ap.add_argument("--con-nexus", action="store_true", help="Consulta Nexus por cada cédula V/E y prioriza identidad oficial")
    ap.add_argument(
        "--solo-cache-nexus",
        action="store_true",
        help="Con --con-nexus, reutiliza caché existente sin consultar cédulas pendientes",
    )
    ap.add_argument(
        "--concurrency",
        type=int,
        default=DEFAULT_CONCURRENCY,
        help=f"Consultas Nexus simultáneas y por segundo (default calibrado: {DEFAULT_CONCURRENCY})",
    )
    ap.add_argument(
        "--rate",
        type=float,
        default=DEFAULT_RATE,
        help=f"Segundos entre consultas si --concurrency=1 (default {DEFAULT_RATE})",
    )
    ap.add_argument(
        "--circuit-breaker",
        type=int,
        default=DEFAULT_CIRCUIT_BREAKER,
        help=f"Corta tras N fallos consecutivos de infraestructura (default {DEFAULT_CIRCUIT_BREAKER})",
    )
    ap.add_argument(
        "--timeout-nexus",
        type=float,
        default=DEFAULT_NEXUS_TIMEOUT,
        help=f"Timeout por consulta Nexus en segundos (default {DEFAULT_NEXUS_TIMEOUT:g})",
    )
    ap.add_argument("--aplicar", action="store_true", help="Escribe en BD vía censo_importar_lote")
    ap.add_argument(
        "--solo-marcar-siipol",
        action="store_true",
        help="No reimporta filas; solo aplica evidencia SIIPOL sobre registros existentes",
    )
    ap.add_argument(
        "--reconciliar-siipol",
        action="store_true",
        help="Usa Documento como lista autoritativa SIIPOL; no importa personas",
    )
    ap.add_argument(
        "--dry-run",
        action="store_true",
        help="No importa censo; las verificaciones Nexus nuevas sí quedan en nexus_consultas",
    )
    ap.add_argument("--lote", type=int, default=200)
    ap.add_argument("--solo-con-cedula", action="store_true", help="Omite filas sin documento")
    ap.add_argument(
        "--permitir-omisiones",
        action="store_true",
        help="Permite aplicar aunque existan filas inválidas o campamentos sin resolver",
    )
    ap.add_argument(
        "--omitir-firmo-presidente",
        action="store_true",
        help="Fuerza firmo_contra_presidente=false (no importa firmas de referéndum)",
    )
    ap.add_argument("--json-out", type=Path, default=None)
    args = ap.parse_args()

    if not args.archivo.exists():
        raise SystemExit(f"No existe {args.archivo}")
    if not 1 <= args.concurrency <= 10:
        raise SystemExit("--concurrency debe estar entre 1 y 10")
    if args.rate < 0:
        raise SystemExit("--rate no puede ser negativo")
    if args.circuit_breaker < 1:
        raise SystemExit("--circuit-breaker debe ser al menos 1")
    if args.timeout_nexus <= 0:
        raise SystemExit("--timeout-nexus debe ser mayor que 0")
    if args.solo_cache_nexus and not args.con_nexus:
        raise SystemExit("--solo-cache-nexus requiere --con-nexus")
    if args.solo_marcar_siipol and not args.aplicar:
        raise SystemExit("--solo-marcar-siipol requiere --aplicar")
    if args.reconciliar_siipol and args.solo_marcar_siipol:
        raise SystemExit("--reconciliar-siipol y --solo-marcar-siipol son excluyentes")

    dry = args.dry_run or not args.aplicar
    print(f"Excel: leyendo {args.archivo.name}…", file=sys.stderr, flush=True)
    rows, hoja_datos, columnas = leer_filas(args.archivo)
    print(
        f"Excel: hoja «{hoja_datos}» · {len(rows)} filas · {len(columnas)} columnas",
        file=sys.stderr,
        flush=True,
    )
    print("Supabase: autenticando…", file=sys.stderr, flush=True)
    url, anon, gateway = cargar_env()
    jwt = autenticar(url, anon)

    if args.reconciliar_siipol:
        documentos_filas = [
            parse_cedula(pick(row, "cedula", "cédula", "documento", "ci", "doc"))[1]
            for row in rows
        ]
        documentos = sorted({documento for documento in documentos_filas if documento})
        reporte_siipol = {
            "archivo": args.archivo.name,
            "hoja": hoja_datos,
            "filas_leidas": len(rows),
            "documentos_validos": sum(bool(documento) for documento in documentos_filas),
            "documentos_unicos": len(documentos),
            "documentos_duplicados": sum(
                cantidad - 1
                for documento, cantidad in Counter(documentos_filas).items()
                if documento and cantidad > 1
            ),
            "sin_documento": sum(not documento for documento in documentos_filas),
        }
        print(json.dumps(reporte_siipol, ensure_ascii=False, indent=2))
        if dry:
            print("Dry-run: no se modificaron marcas SIIPOL.", file=sys.stderr)
            return 0
        resultado = rpc(
            url,
            anon,
            jwt,
            "censo_reconciliar_siipol",
            {"p_documentos": documentos, "p_fuente": args.archivo.name},
        )
        print(json.dumps(resultado, ensure_ascii=False, indent=2))
        return 0

    print("Supabase: cargando campamentos…", file=sys.stderr, flush=True)
    centros = listar_centros(url, anon, jwt)

    preparadas: list[dict[str, Any]] = []
    ok: list[dict[str, Any]] = []
    errores: list[dict[str, Any]] = []
    nexus_errores: list[dict[str, Any]] = []
    conteos = {
        "con_cedula": 0,
        "sin_cedula": 0,
        "nexus_cache": 0,
        "nexus_ya_verificadas_unicas": 0,
        "nexus_consultadas": 0,
        "nexus_verificadas_nuevas": 0,
        "nexus_ok": 0,
        "nexus_error": 0,
        "nexus_omitidas_circuit_breaker": 0,
        "nexus_omitidas_solo_cache": 0,
        "solicitados": 0,
        "registro_policial": 0,
        "firmo_contra_presidente": 0,
        "verificados_siipol": 0,
    }

    for row in rows:
        _, documento_entrada = parse_cedula(pick(row, "cedula", "cédula", "documento", "ci", "doc"))
        if documento_entrada:
            conteos["con_cedula"] += 1
        else:
            conteos["sin_cedula"] += 1
        payload, error = fila_a_payload(row, centros, args.centro_id, args.col_centro)
        if error:
            errores.append(error)
            continue
        if payload is None:
            errores.append({"error": "fila_incompleta", "row": row})
            continue

        if not payload.get("documento") and args.solo_con_cedula:
            continue
        preparadas.append(payload)

    cache_persistente: dict[tuple[str, str], dict[str, Any]] = {}
    fichas_nuevas: dict[tuple[str, str], dict[str, Any]] = {}
    candidatos: set[tuple[str, str]] = set()
    if args.con_nexus:
        for payload in preparadas:
            tipo_doc = texto(payload.get("tipo_doc")).upper()
            documento = texto(payload.get("documento"))
            if tipo_doc in LETRAS_NEXUS and documento:
                candidatos.add((tipo_doc, documento))

        print("Nexus: cargando verificaciones existentes…", file=sys.stderr, flush=True)
        cache_persistente = cargar_cache_nexus(url, anon, jwt)
        pendientes = sorted(candidatos - cache_persistente.keys())
        conteos["nexus_ya_verificadas_unicas"] = len(candidatos) - len(pendientes)
        pendientes_totales = len(pendientes)
        print(
            f"Nexus: {len(candidatos)} cédulas únicas · "
            f"{len(candidatos) - len(pendientes)} ya verificadas · "
            f"{len(pendientes)} pendientes · concurrencia {args.concurrency}",
            file=sys.stderr,
        )
        if args.solo_cache_nexus:
            conteos["nexus_omitidas_solo_cache"] = pendientes_totales
            pendientes = []
            print(
                f"Nexus: modo solo caché · {pendientes_totales} pendientes usarán datos del Excel",
                file=sys.stderr,
                flush=True,
            )

        if pendientes:
            fichas_nuevas, nexus_errores, omitidas = consultar_nexus_concurrente(
                gateway,
                jwt,
                pendientes,
                args.concurrency,
                args.rate,
                args.circuit_breaker,
                args.timeout_nexus,
            )
            conteos["nexus_consultadas"] = len(fichas_nuevas) + len(nexus_errores)
            conteos["nexus_verificadas_nuevas"] = len(fichas_nuevas)
            conteos["nexus_omitidas_circuit_breaker"] = omitidas
            # El dry-run evita importar censo, pero la verificación Nexus se
            # persiste: la confirmación posterior nunca repite estas consultas.
            guardar_cache_nexus(url, anon, jwt, fichas_nuevas)

        fichas_disponibles = {**cache_persistente, **fichas_nuevas}
        for payload in preparadas:
            clave = (
                texto(payload.get("tipo_doc")).upper(),
                texto(payload.get("documento")),
            )
            ficha = fichas_disponibles.get(clave)
            if ficha is not None:
                fuente_nexus = "cache" if clave in cache_persistente else "nexus"
                aplicar_nexus(payload, ficha, fuente_nexus)
                conteos["nexus_ok"] += 1
                if clave in cache_persistente:
                    conteos["nexus_cache"] += 1
            elif clave[0] in LETRAS_NEXUS and clave[1]:
                conteos["nexus_error"] += 1

    firmo_omitidos = 0
    for payload in preparadas:
        if not payload.get("primer_nombre") or not payload.get("primer_apellido"):
            errores.append({"error": "sin_nombre", "documento": payload.get("documento")})
            continue

        if args.omitir_firmo_presidente and payload.get("firmo_contra_presidente"):
            firmo_omitidos += 1
            payload["firmo_contra_presidente"] = False

        if payload.get("solicitado"):
            conteos["solicitados"] += 1
        if payload.get("registro_policial"):
            conteos["registro_policial"] += 1
        if payload.get("firmo_contra_presidente"):
            conteos["firmo_contra_presidente"] += 1
        if payload.get("verificado_siipol"):
            conteos["verificados_siipol"] += 1
        ok.append(payload)

    ok.sort(key=lambda f: (0 if f.get("documento") else 1, f.get("primer_apellido") or ""))
    documentos = [texto(f.get("documento")) for f in ok if f.get("documento")]
    documentos_repetidos = sum(cantidad - 1 for cantidad in Counter(documentos).values() if cantidad > 1)
    errores_por_tipo = Counter(texto(error.get("error")) or "desconocido" for error in errores)
    centros_con_error = Counter(
        texto(error.get("nombre_centro_raw"))
        for error in errores
        if error.get("nombre_centro_raw")
    )
    sensibles_ignoradas = [
        columna
        for columna in columnas
        if key(columna) in {"milita oposicion"}
    ]
    reporte = {
        "archivo": args.archivo.name,
        "hoja": hoja_datos,
        "filas_leidas": len(rows),
        "listas": len(ok),
        **conteos,
        "firmo_omitidos": firmo_omitidos if args.omitir_firmo_presidente else 0,
        "documentos_repetidos": documentos_repetidos,
        "errores_match": len(errores),
        "errores_por_tipo": dict(sorted(errores_por_tipo.items())),
        "centros_con_error": dict(sorted(centros_con_error.items())),
        "columnas_sensibles_ignoradas": sensibles_ignoradas,
        "nexus_errores_muestra": [
            {"tipo_doc": error.get("tipo_doc"), "status": error.get("status")}
            for error in nexus_errores[:10]
        ],
    }
    print(json.dumps(reporte, ensure_ascii=False, indent=2))

    if args.json_out:
        args.json_out.write_text(
            json.dumps({"filas": ok, "errores": errores, "nexus_errores": nexus_errores}, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(f"JSON escrito: {args.json_out}", file=sys.stderr)

    if dry:
        detalle_nexus = (
            " Verificaciones Nexus nuevas quedaron guardadas en nexus_consultas; "
            "--aplicar las reutilizará."
            if args.con_nexus and not args.solo_cache_nexus
            else ""
        )
        print(f"Dry-run: no se importó el censo.{detalle_nexus}", file=sys.stderr)
        return 0
    if errores and not args.permitir_omisiones:
        raise SystemExit(
            f"Importación cancelada: {len(errores)} filas no están listas. "
            "Corrija los errores o use --permitir-omisiones tras aprobar una importación parcial."
        )

    insertados = actualizados = omitidos = marcados_siipol = 0
    errores_rpc: list[Any] = []
    for i in range(0, len(ok), args.lote):
        chunk = ok[i : i + args.lote]
        result: object = {"modo": "solo_marcar_siipol"}
        if not args.solo_marcar_siipol:
            result = rpc(
                url,
                anon,
                jwt,
                "censo_importar_lote",
                {"p_filas": chunk, "p_meta": {"fuente_archivo": args.archivo.name}},
            )
            if isinstance(result, dict):
                insertados += int(result.get("insertados") or 0)
                actualizados += int(result.get("actualizados") or 0)
                omitidos += int(result.get("omitidos") or 0)
                err = result.get("errores") or []
                if isinstance(err, list):
                    errores_rpc.extend(err)
        if args.solo_marcar_siipol:
            resultado_siipol = rpc(
                url,
                anon,
                jwt,
                "censo_marcar_siipol_lote",
                {"p_filas": chunk, "p_fuente": args.archivo.name},
            )
            if isinstance(resultado_siipol, dict):
                marcados_siipol += int(resultado_siipol.get("marcados_siipol") or 0)
        print(f"Lote {i // args.lote + 1}: {result}", file=sys.stderr)

    print(
        json.dumps(
            {
                "insertados": insertados,
                "actualizados": actualizados,
                "omitidos": omitidos,
                "marcados_siipol": marcados_siipol,
                "errores_rpc": errores_rpc[:30],
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
